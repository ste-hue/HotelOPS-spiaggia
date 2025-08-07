#!/usr/bin/env python3
"""
Script per estrarre tutte le formule dal foglio MASTER del file Excel
usando openpyxl per analizzare la struttura e la logica di calcolo
"""

import openpyxl
from openpyxl import load_workbook
import json
import csv
from datetime import datetime
import sys
import os

def extract_excel_formulas(file_path, sheet_name="NUOVO MASTER"):
    """
    Estrae tutte le formule da un foglio Excel specifico
    
    Args:
        file_path (str): Percorso al file Excel
        sheet_name (str): Nome del foglio da analizzare
    
    Returns:
        dict: Dizionario con tutte le informazioni estratte
    """
    try:
        # Carica il workbook
        print(f"📂 Caricamento file: {file_path}")
        workbook = load_workbook(filename=file_path, data_only=False)
        
        # Verifica se il foglio esiste
        if sheet_name not in workbook.sheetnames:
            print(f"❌ Foglio '{sheet_name}' non trovato!")
            print(f"📋 Fogli disponibili: {workbook.sheetnames}")
            return None
        
        # Seleziona il foglio
        worksheet = workbook[sheet_name]
        print(f"✅ Foglio '{sheet_name}' caricato con successo")
        
        # Struttura dati per le informazioni estratte
        extracted_data = {
            "file_info": {
                "filename": os.path.basename(file_path),
                "sheet_name": sheet_name,
                "extraction_date": datetime.now().isoformat(),
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column
            },
            "formulas": [],
            "values": [],
            "merged_cells": [],
            "formatted_cells": [],
            "structure_analysis": {}
        }
        
        print(f"📊 Dimensioni foglio: {worksheet.max_row} righe × {worksheet.max_column} colonne")
        
        # Estrai celle unite
        for merged_range in worksheet.merged_cells.ranges:
            extracted_data["merged_cells"].append({
                "range": str(merged_range),
                "start_cell": merged_range.min_col, 
                "start_row": merged_range.min_row,
                "end_col": merged_range.max_col,
                "end_row": merged_range.max_row
            })
        
        # Analizza ogni cella
        formula_count = 0
        value_count = 0
        
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_address = cell.coordinate
                
                # Estrai formule
                if cell.data_type == 'f':  # Formula
                    formula_count += 1
                    extracted_data["formulas"].append({
                        "cell": cell_address,
                        "row": row,
                        "column": col,
                        "formula": cell.value,
                        "display_value": cell.displayed_value if hasattr(cell, 'displayed_value') else None
                    })
                
                # Estrai valori (non formule)
                elif cell.value is not None:
                    value_count += 1
                    extracted_data["values"].append({
                        "cell": cell_address,
                        "row": row,
                        "column": col,
                        "value": cell.value,
                        "data_type": cell.data_type,
                        "number_format": cell.number_format if cell.number_format != 'General' else None
                    })
                
                # Estrai formattazione interessante
                if (cell.fill.start_color.index != '00000000' or 
                    cell.font.bold or 
                    cell.number_format != 'General'):
                    extracted_data["formatted_cells"].append({
                        "cell": cell_address,
                        "row": row,
                        "column": col,
                        "background_color": cell.fill.start_color.index,
                        "font_bold": cell.font.bold,
                        "number_format": cell.number_format,
                        "value": str(cell.value)[:50] if cell.value else None  # Primi 50 char
                    })
        
        print(f"🧮 Formule trovate: {formula_count}")
        print(f"📝 Valori trovati: {value_count}")
        print(f"🎨 Celle formattate: {len(extracted_data['formatted_cells'])}")
        print(f"🔗 Celle unite: {len(extracted_data['merged_cells'])}")
        
        # Analisi strutturale
        extracted_data["structure_analysis"] = analyze_structure(extracted_data)
        
        return extracted_data
        
    except FileNotFoundError:
        print(f"❌ File non trovato: {file_path}")
        return None
    except Exception as e:
        print(f"❌ Errore durante l'estrazione: {str(e)}")
        return None

def analyze_structure(data):
    """
    Analizza la struttura del foglio per identificare sezioni logiche
    """
    analysis = {
        "sections_identified": [],
        "formula_patterns": {},
        "calculation_chains": []
    }
    
    # Identifica pattern nelle formule
    formula_types = {}
    for formula_info in data["formulas"]:
        formula = formula_info["formula"]
        if formula:
            if formula.startswith("=SUM"):
                formula_types["SUM"] = formula_types.get("SUM", 0) + 1
            elif "*" in formula:
                formula_types["MULTIPLICATION"] = formula_types.get("MULTIPLICATION", 0) + 1
            elif "+" in formula:
                formula_types["ADDITION"] = formula_types.get("ADDITION", 0) + 1
    
    analysis["formula_patterns"] = formula_types
    
    # Identifica sezioni basate su celle unite e formattazione
    sections = []
    for merged in data["merged_cells"]:
        sections.append({
            "type": "header_section",
            "range": merged["range"],
            "location": f"Row {merged['start_row']}"
        })
    
    analysis["sections_identified"] = sections
    
    return analysis

def save_results(data, output_dir="./output"):
    """
    Salva i risultati in diversi formati
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 1. Salva tutto in JSON
    json_file = f"{output_dir}/formulas_analysis_{timestamp}.json"
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    print(f"💾 Analisi completa salvata in: {json_file}")
    
    # 2. Salva solo le formule in CSV
    csv_file = f"{output_dir}/formulas_only_{timestamp}.csv"
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Cell', 'Row', 'Column', 'Formula', 'Display_Value'])
        for formula in data["formulas"]:
            writer.writerow([
                formula['cell'], 
                formula['row'], 
                formula['column'], 
                formula['formula'], 
                formula.get('display_value', '')
            ])
    print(f"📊 Formule salvate in CSV: {csv_file}")
    
    # 3. Salva report leggibile
    report_file = f"{output_dir}/analysis_report_{timestamp}.txt"
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("="*60 + "\n")
        f.write("ANALISI FORMULE EXCEL - MASTER SHEET\n")
        f.write("="*60 + "\n\n")
        
        f.write(f"File: {data['file_info']['filename']}\n")
        f.write(f"Foglio: {data['file_info']['sheet_name']}\n")
        f.write(f"Data estrazione: {data['file_info']['extraction_date']}\n")
        f.write(f"Dimensioni: {data['file_info']['max_row']}x{data['file_info']['max_column']}\n\n")
        
        f.write("FORMULE TROVATE:\n")
        f.write("-"*40 + "\n")
        for formula in data["formulas"]:
            f.write(f"Cella {formula['cell']}: {formula['formula']}\n")
        
        f.write(f"\nPATTERN FORMULE:\n")
        f.write("-"*40 + "\n")
        for pattern, count in data["structure_analysis"]["formula_patterns"].items():
            f.write(f"{pattern}: {count} occorrenze\n")
        
        f.write(f"\nSEZIONI IDENTIFICATE:\n")
        f.write("-"*40 + "\n")
        for section in data["structure_analysis"]["sections_identified"]:
            f.write(f"{section['type']}: {section['range']} (Row {section['location']})\n")
    
    print(f"📋 Report leggibile salvato in: {report_file}")

def print_formula_summary(data):
    """
    Stampa un riassunto delle formule trovate
    """
    print("\n" + "="*60)
    print("🔍 RIASSUNTO FORMULE ESTRATTE")
    print("="*60)
    
    if not data["formulas"]:
        print("❌ Nessuna formula trovata!")
        return
    
    print(f"\n📊 Totale formule: {len(data['formulas'])}")
    print("\n🧮 FORMULE TROVATE:")
    print("-"*40)
    
    for i, formula in enumerate(data["formulas"], 1):
        print(f"{i:2d}. Cella {formula['cell']:<6} → {formula['formula']}")
    
    print(f"\n📈 PATTERN ANALIZZATI:")
    print("-"*40)
    for pattern, count in data["structure_analysis"]["formula_patterns"].items():
        print(f"• {pattern:<15}: {count:2d} volte")

def main():
    """
    Funzione principale
    """
    print("🏖️ ESTRATTORE FORMULE EXCEL - MASTER SHEET")
    print("="*50)
    
    # Percorso del file Excel
    excel_file = "/Users/stefanodellapietra/Desktop/Projects/Companies/INTUR/INTUR_development/HotelOPS/modules/spiaggia/chiusura/cassa_spiaggia.xlsx"
    
    # Verifica se il file esiste
    if not os.path.exists(excel_file):
        print(f"❌ File non trovato: {excel_file}")
        print("💡 Assicurati che il percorso sia corretto")
        return
    
    # Estrai le formule
    extracted_data = extract_excel_formulas(excel_file, "NUOVO MASTER")
    
    if extracted_data is None:
        print("❌ Estrazione fallita!")
        return
    
    # Mostra riassunto
    print_formula_summary(extracted_data)
    
    # Salva i risultati
    save_results(extracted_data)
    
    print(f"\n✅ Estrazione completata con successo!")
    print(f"📁 Controlla la cartella './output' per i file generati")

if __name__ == "__main__":
    main()
